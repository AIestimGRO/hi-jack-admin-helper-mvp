from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.clients import upsert_client
from app.services.phone import normalize_phone


FIELD_LABELS = {
    "app_user_id": "ID",
    "telegram_id": "Telegram ID",
    "referrer_app_user_id": "Referrer ID",
    "nickname": "Nickname",
    "username": "Username",
    "first_name": "First name",
    "phone_raw": "Phone",
}

ALIASES = {
    "app_user_id": {"id", "user id", "app user id", "userid"},
    "telegram_id": {"telegram id", "telegram_id", "tg id", "tg_id"},
    "referrer_app_user_id": {"referrer id", "referrer_id", "ref id", "реферер id"},
    "nickname": {"nickname", "nick", "никнейм"},
    "username": {"username", "user name", "telegram username", "юзернейм"},
    "first_name": {"first name", "firstname", "name", "имя"},
    "phone_raw": {"phone", "phone number", "telephone", "телефон", "номер телефона"},
}


def _normalized_header(value: Any) -> str:
    return " ".join(str(value or "").replace("_", " ").strip().lower().split())


def detect_mapping(headers: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    normalized = {_normalized_header(header): header for header in headers}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if _normalized_header(alias) in normalized:
                result[field] = normalized[_normalized_header(alias)]
                break
    return result


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("unsupported_csv_encoding")


def read_tabular(path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration:
            return [], []
        headers = [str(value or "").strip() for value in raw_headers]
        rows = [dict(zip(headers, values)) for values in iterator if any(value not in (None, "") for value in values)]
        workbook.close()
        return headers, rows

    text = _decode_csv(path.read_bytes())
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [str(value or "").strip() for value in (reader.fieldnames or [])]
    rows = []
    for raw_row in reader:
        if not any(value not in (None, "") for value in raw_row.values()):
            continue
        rows.append({str(key or "").strip(): value for key, value in raw_row.items() if key is not None})
    return headers, rows


def mapped_row(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    return {field: row.get(column) for field, column in mapping.items() if column}


def import_rows(conn, rows: list[dict[str, Any]], mapping: dict[str, str]) -> dict[str, int]:
    stats = {
        "total": len(rows),
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "phone_errors": 0,
        "duplicates": 0,
    }
    seen_phones: set[str] = set()
    seen_app_ids: set[str] = set()
    for row in rows:
        values = mapped_row(row, mapping)
        phone_raw = values.get("phone_raw")
        phone_local = normalize_phone(phone_raw)
        app_id = str(values.get("app_user_id") or "").strip()
        if phone_raw and not phone_local:
            stats["phone_errors"] += 1
        if phone_local in seen_phones or (app_id and app_id in seen_app_ids):
            stats["duplicates"] += 1
        if phone_local:
            seen_phones.add(phone_local)
        if app_id:
            seen_app_ids.add(app_id)
        if not phone_local and not app_id:
            stats["skipped"] += 1
            continue
        try:
            _, action = upsert_client(conn, values, source="hijack_app")
            stats[action] += 1
        except ValueError:
            stats["skipped"] += 1
    return stats

