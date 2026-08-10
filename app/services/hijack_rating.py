from __future__ import annotations

import io
import json
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import load_workbook

from app.services.phone import normalize_phone

HIJACK_TITLE_CONDITIONS: dict[str, str] = {
    "hijack_year_rating": "HI, JACK! · рейтинг за год",
    "hijack_month_rating": "HI, JACK! · рейтинг за месяц",
    "hijack_latest_rating": "HI, JACK! · рейтинг последнего турнира",
    "hijack_year_kills": "HI, JACK! · киллы за год",
    "hijack_month_kills": "HI, JACK! · киллы за месяц",
    "hijack_latest_kills": "HI, JACK! · киллы последнего турнира",
    "hijack_tournaments_played": "HI, JACK! · сыграно турниров",
    "hijack_top3_finishes": "HI, JACK! · финиши в топ-3",
    "hijack_wins": "HI, JACK! · победы в турнирах",
    "hijack_best_rating": "HI, JACK! · лучший рейтинг за один турнир",
}

MAX_IMPORT_BYTES = 12 * 1024 * 1024
MAX_REFERRAL_TREE_DEPTH = 20


def ensure_hijack_rating_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS hi_jack_rating_imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tournament_name TEXT NOT NULL,
            tournament_date TEXT NOT NULL,
            source_filename TEXT NOT NULL DEFAULT '',
            total_rows INTEGER NOT NULL DEFAULT 0,
            matched_rows INTEGER NOT NULL DEFAULT 0,
            unmatched_rows INTEGER NOT NULL DEFAULT 0,
            invalid_rows INTEGER NOT NULL DEFAULT 0,
            uploaded_by_admin_id INTEGER REFERENCES admins(id) ON DELETE SET NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS hi_jack_rating_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL
                REFERENCES hi_jack_rating_imports(id) ON DELETE CASCADE,
            client_id INTEGER REFERENCES clients(id) ON DELETE SET NULL,
            phone_local TEXT,
            phone_raw TEXT NOT NULL DEFAULT '',
            rating_points REAL NOT NULL DEFAULT 0,
            kills INTEGER NOT NULL DEFAULT 0,
            source_row INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(import_id, source_row)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_hi_jack_rating_imports_date
        ON hi_jack_rating_imports(tournament_date, id)
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_hi_jack_rating_entries_client
        ON hi_jack_rating_entries(client_id, import_id)
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_hi_jack_rating_entries_phone
        ON hi_jack_rating_entries(phone_local, import_id)
        """
    )


def _clean_header(value: Any) -> str:
    text = str(value or "").replace("Ё", "Е").replace("ё", "е").strip().upper()
    text = re.sub(r"[._:;]+", " ", text)
    return " ".join(text.split())


def _number(value: Any, *, integer: bool = False) -> float | int:
    if value is None or value == "":
        return 0 if integer else 0.0
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
        text = re.sub(r"[^0-9.\-]", "", text)
        if not text or text in {"-", ".", "-."}:
            return 0 if integer else 0.0
        number = float(text)
    if integer:
        return max(0, int(round(number)))
    return max(0.0, number)


def parse_hijack_rating_workbook(data: bytes) -> list[dict[str, Any]]:
    if not data:
        raise ValueError("Выберите Excel-файл")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("Excel-файл слишком большой")
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось открыть Excel-файл") from exc
    try:
        sheet = workbook.active
        header_row = None
        column_map: dict[str, int] = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number > 80:
                break
            headers = {_clean_header(value): index for index, value in enumerate(row) if value is not None}
            phone_index = next((headers[key] for key in ("PHONE", "ТЕЛЕФОН") if key in headers), None)
            rating_index = next(
                (headers[key] for key in ("ИГР РЕЙТ", "ИГР РЕЙТИНГ", "ИГРОВОЙ РЕЙТИНГ") if key in headers),
                None,
            )
            kills_index = next(
                (headers[key] for key in ("ИГР КИЛ", "ИГР КИЛЛ", "ИГР КИЛЫ", "ИГР КИЛЛЫ") if key in headers),
                None,
            )
            if phone_index is not None and rating_index is not None and kills_index is not None:
                header_row = row_number
                column_map = {"phone": phone_index, "rating": rating_index, "kills": kills_index}
                break
        if header_row is None:
            raise ValueError("Не найдены обязательные столбцы Phone, ИГР Рейт и ИГР Кил")

        result: list[dict[str, Any]] = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number <= header_row:
                continue
            values = list(row)
            max_index = max(column_map.values())
            if len(values) <= max_index:
                values.extend([None] * (max_index + 1 - len(values)))
            phone_raw = values[column_map["phone"]]
            rating_raw = values[column_map["rating"]]
            kills_raw = values[column_map["kills"]]
            if phone_raw in {None, ""} and rating_raw in {None, ""} and kills_raw in {None, ""}:
                continue
            result.append(
                {
                    "source_row": row_number,
                    "phone_raw": str(phone_raw or "").strip(),
                    "phone_local": normalize_phone(phone_raw),
                    "rating_points": float(_number(rating_raw)),
                    "kills": int(_number(kills_raw, integer=True)),
                }
            )
        if not result:
            raise ValueError("В Excel-файле нет строк рейтинга")
        return result
    finally:
        workbook.close()


def import_hijack_rating(
    conn: sqlite3.Connection,
    *,
    tournament_name: str,
    tournament_date: date,
    source_filename: str,
    rows: list[dict[str, Any]],
    admin_id: int | None,
) -> dict[str, Any]:
    ensure_hijack_rating_schema(conn)
    name = " ".join(str(tournament_name or "").split())[:160]
    if not name:
        raise ValueError("Укажите название турнира")
    cursor = conn.execute(
        """
        INSERT INTO hi_jack_rating_imports(
            tournament_name,tournament_date,source_filename,uploaded_by_admin_id
        ) VALUES (?,?,?,?)
        """,
        (name, tournament_date.isoformat(), str(source_filename or "")[:250], admin_id),
    )
    import_id = int(cursor.lastrowid)
    matched = 0
    unmatched = 0
    invalid = 0
    inserted = 0
    for item in rows:
        phone_local = item.get("phone_local")
        client_id = None
        if phone_local:
            client = conn.execute(
                "SELECT id FROM clients WHERE phone_local=? LIMIT 1", (phone_local,)
            ).fetchone()
            if client:
                client_id = int(client["id"])
                matched += 1
            else:
                unmatched += 1
        else:
            invalid += 1
        conn.execute(
            """
            INSERT INTO hi_jack_rating_entries(
                import_id,client_id,phone_local,phone_raw,rating_points,kills,source_row
            ) VALUES (?,?,?,?,?,?,?)
            """,
            (
                import_id,
                client_id,
                phone_local,
                str(item.get("phone_raw") or "")[:80],
                float(item.get("rating_points") or 0),
                int(item.get("kills") or 0),
                int(item.get("source_row") or 0),
            ),
        )
        inserted += 1
    conn.execute(
        """
        UPDATE hi_jack_rating_imports
        SET total_rows=?,matched_rows=?,unmatched_rows=?,invalid_rows=?
        WHERE id=?
        """,
        (inserted, matched, unmatched, invalid, import_id),
    )
    _expire_previous_latest_titles(conn)
    return {
        "import_id": import_id,
        "tournament_name": name,
        "tournament_date": tournament_date.isoformat(),
        "total_rows": inserted,
        "matched_rows": matched,
        "unmatched_rows": unmatched,
        "invalid_rows": invalid,
    }


def list_hijack_imports(conn: sqlite3.Connection, *, limit: int = 30) -> list[dict[str, Any]]:
    ensure_hijack_rating_schema(conn)
    rows = conn.execute(
        """
        SELECT * FROM hi_jack_rating_imports
        ORDER BY created_at DESC,id DESC LIMIT ?
        """,
        (max(1, min(int(limit), 100)),),
    ).fetchall()
    return [dict(row) for row in rows]


def _display_name(row: sqlite3.Row | dict[str, Any]) -> str:
    for key in ("nickname", "first_name", "username"):
        value = str(row[key] or "").strip() if key in row.keys() else ""
        if value:
            return f"@{value.lstrip('@')}" if key == "username" else value
    client_id = int(row["client_id"] if "client_id" in row.keys() else row["id"])
    return f"HJ #{client_id}"


def _period_aggregate(
    conn: sqlite3.Connection,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    import_id: int | None = None,
) -> list[dict[str, Any]]:
    where = ["e.client_id IS NOT NULL"]
    params: list[Any] = []
    if import_id is not None:
        where.append("e.import_id=?")
        params.append(import_id)
    if start_date is not None:
        where.append("i.tournament_date>=?")
        params.append(start_date.isoformat())
    if end_date is not None:
        where.append("i.tournament_date<?")
        params.append(end_date.isoformat())
    rows = conn.execute(
        f"""
        SELECT e.client_id,c.nickname,c.first_name,c.username,
               SUM(e.rating_points) AS points,
               SUM(e.kills) AS kills,
               COUNT(DISTINCT e.import_id) AS tournaments
        FROM hi_jack_rating_entries e
        JOIN hi_jack_rating_imports i ON i.id=e.import_id
        JOIN clients c ON c.id=e.client_id
        WHERE {' AND '.join(where)}
        GROUP BY e.client_id,c.nickname,c.first_name,c.username
        """,
        params,
    ).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        item["display_name"] = _display_name(row)
        points = float(item["points"] or 0)
        item["points"] = int(points) if points.is_integer() else round(points, 1)
        item["kills"] = int(item["kills"] or 0)
        item["tournaments"] = int(item["tournaments"] or 0)
        result.append(item)
    result.sort(
        key=lambda item: (
            -float(item["points"]),
            -int(item["kills"]),
            -int(item["tournaments"]),
            int(item["client_id"]),
        )
    )
    previous_key = None
    place = 0
    for index, item in enumerate(result, start=1):
        current_key = (float(item["points"]), int(item["kills"]))
        if current_key != previous_key:
            place = index
            previous_key = current_key
        item["place"] = place
    return result


def _member_row(rows: list[dict[str, Any]], client_id: int) -> dict[str, Any] | None:
    return next((row for row in rows if int(row["client_id"]) == int(client_id)), None)


def hijack_rating_payload(
    conn: sqlite3.Connection,
    *,
    client_id: int,
    today: date | None = None,
) -> dict[str, Any]:
    ensure_hijack_rating_schema(conn)
    today = today or date.today()
    year_start = date(today.year, 1, 1)
    year_end = date(today.year + 1, 1, 1)
    month_start = today.replace(day=1)
    if today.month == 12:
        month_end = date(today.year + 1, 1, 1)
    else:
        month_end = date(today.year, today.month + 1, 1)
    latest = conn.execute(
        "SELECT * FROM hi_jack_rating_imports ORDER BY created_at DESC,id DESC LIMIT 1"
    ).fetchone()
    year_rows = _period_aggregate(conn, start_date=year_start, end_date=year_end)
    month_rows = _period_aggregate(conn, start_date=month_start, end_date=month_end)
    latest_rows = _period_aggregate(conn, import_id=int(latest["id"])) if latest else []
    return {
        "has_imports": bool(latest),
        "year": {
            "label": str(today.year),
            "rows": year_rows,
            "me": _member_row(year_rows, client_id),
        },
        "month": {
            "label": month_start.strftime("%m.%Y"),
            "rows": month_rows,
            "me": _member_row(month_rows, client_id),
        },
        "latest": {
            "tournament_name": str(latest["tournament_name"]) if latest else "",
            "tournament_date": str(latest["tournament_date"]) if latest else "",
            "rows": latest_rows,
            "me": _member_row(latest_rows, client_id),
        },
    }


def _client_points(
    conn: sqlite3.Connection,
    *,
    client_id: int,
    start_date: date | None = None,
    end_date: date | None = None,
    import_id: int | None = None,
) -> tuple[float, int]:
    where = ["e.client_id=?"]
    params: list[Any] = [client_id]
    if import_id is not None:
        where.append("e.import_id=?")
        params.append(import_id)
    if start_date is not None:
        where.append("i.tournament_date>=?")
        params.append(start_date.isoformat())
    if end_date is not None:
        where.append("i.tournament_date<?")
        params.append(end_date.isoformat())
    row = conn.execute(
        f"""
        SELECT COALESCE(SUM(e.rating_points),0) AS points,
               COALESCE(SUM(e.kills),0) AS kills
        FROM hi_jack_rating_entries e
        JOIN hi_jack_rating_imports i ON i.id=e.import_id
        WHERE {' AND '.join(where)}
        """,
        params,
    ).fetchone()
    return float(row["points"] or 0), int(row["kills"] or 0)


def hijack_member_metrics(
    conn: sqlite3.Connection,
    *,
    client_id: int,
    today: date | None = None,
) -> dict[str, int | float]:
    ensure_hijack_rating_schema(conn)
    today = today or date.today()
    year_start = date(today.year, 1, 1)
    year_end = date(today.year + 1, 1, 1)
    month_start = today.replace(day=1)
    month_end = date(today.year + 1, 1, 1) if today.month == 12 else date(today.year, today.month + 1, 1)
    latest = conn.execute(
        "SELECT id FROM hi_jack_rating_imports ORDER BY created_at DESC,id DESC LIMIT 1"
    ).fetchone()
    year_points, year_kills = _client_points(
        conn, client_id=client_id, start_date=year_start, end_date=year_end
    )
    month_points, month_kills = _client_points(
        conn, client_id=client_id, start_date=month_start, end_date=month_end
    )
    latest_points, latest_kills = (
        _client_points(conn, client_id=client_id, import_id=int(latest["id"]))
        if latest
        else (0.0, 0)
    )
    tournaments = int(
        conn.execute(
            "SELECT COUNT(DISTINCT import_id) FROM hi_jack_rating_entries WHERE client_id=?",
            (client_id,),
        ).fetchone()[0]
        or 0
    )
    best_rating = float(
        conn.execute(
            """
            SELECT COALESCE(MAX(points),0) FROM (
                SELECT SUM(rating_points) AS points
                FROM hi_jack_rating_entries
                WHERE client_id=? GROUP BY import_id
            )
            """,
            (client_id,),
        ).fetchone()[0]
        or 0
    )
    placements = conn.execute(
        """
        WITH totals AS (
            SELECT import_id,client_id,SUM(rating_points) AS points,SUM(kills) AS kills
            FROM hi_jack_rating_entries
            WHERE client_id IS NOT NULL
            GROUP BY import_id,client_id
        ), ranked AS (
            SELECT import_id,client_id,
                   RANK() OVER (PARTITION BY import_id ORDER BY points DESC,kills DESC,client_id) AS place
            FROM totals
        )
        SELECT
            COALESCE(SUM(CASE WHEN place<=3 THEN 1 ELSE 0 END),0) AS top3,
            COALESCE(SUM(CASE WHEN place=1 THEN 1 ELSE 0 END),0) AS wins
        FROM ranked WHERE client_id=?
        """,
        (client_id,),
    ).fetchone()
    return {
        "hijack_year_rating": int(year_points) if year_points.is_integer() else round(year_points, 1),
        "hijack_month_rating": int(month_points) if month_points.is_integer() else round(month_points, 1),
        "hijack_latest_rating": int(latest_points) if latest_points.is_integer() else round(latest_points, 1),
        "hijack_year_kills": year_kills,
        "hijack_month_kills": month_kills,
        "hijack_latest_kills": latest_kills,
        "hijack_tournaments_played": tournaments,
        "hijack_top3_finishes": int(placements["top3"] or 0),
        "hijack_wins": int(placements["wins"] or 0),
        "hijack_best_rating": int(best_rating) if best_rating.is_integer() else round(best_rating, 1),
    }


def _temporary_bounds(condition_code: str, today: date, latest_import_id: int | None) -> tuple[str, datetime, datetime]:
    start = datetime.combine(today, datetime.min.time(), tzinfo=timezone.utc)
    if "_month_" in condition_code:
        start_date = today.replace(day=1)
        end_date = date(today.year + 1, 1, 1) if today.month == 12 else date(today.year, today.month + 1, 1)
        return (
            start_date.strftime("%Y-%m"),
            datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc),
            datetime.combine(end_date, datetime.min.time(), tzinfo=timezone.utc),
        )
    if "_latest_" in condition_code:
        return (
            f"latest-{latest_import_id or 0}",
            start,
            datetime(9999, 12, 31, tzinfo=timezone.utc),
        )
    start_date = date(today.year, 1, 1)
    end_date = date(today.year + 1, 1, 1)
    return (
        str(today.year),
        datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc),
        datetime.combine(end_date, datetime.min.time(), tzinfo=timezone.utc),
    )


def _expire_previous_latest_titles(conn: sqlite3.Connection) -> None:
    now_iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
    conn.execute(
        """
        UPDATE member_titles
        SET expires_at=?
        WHERE temporary_period_id IS NOT NULL
          AND title_definition_id IN (
              SELECT id FROM title_definitions WHERE condition_code LIKE 'hijack_latest_%'
          )
          AND (expires_at IS NULL OR expires_at>?)
        """,
        (now_iso, now_iso),
    )


def refresh_hijack_engagement(
    conn: sqlite3.Connection,
    *,
    client_id: int,
    today: date | None = None,
) -> dict[str, int | float]:
    ensure_hijack_rating_schema(conn)
    today = today or date.today()
    metrics = hijack_member_metrics(conn, client_id=client_id, today=today)
    latest = conn.execute(
        "SELECT id FROM hi_jack_rating_imports ORDER BY created_at DESC,id DESC LIMIT 1"
    ).fetchone()
    latest_import_id = int(latest["id"]) if latest else None

    from app.services.jackside_engagement import (  # local import avoids startup cycles
        _grant_configured_material_reward,
        _notify,
    )

    achievements = conn.execute(
        "SELECT * FROM achievement_definitions WHERE is_enabled=1 AND condition_code LIKE 'hijack_%'"
    ).fetchall()
    for definition in achievements:
        value = float(metrics.get(str(definition["condition_code"]), 0) or 0)
        if value < float(definition["threshold"]):
            continue
        cursor = conn.execute(
            """
            INSERT OR IGNORE INTO member_achievements(client_id,achievement_id,source_json)
            VALUES (?,?,?)
            """,
            (
                client_id,
                int(definition["id"]),
                json.dumps({"value": value, "threshold": int(definition["threshold"]), "source": "hi_jack"}, ensure_ascii=False),
            ),
        )
        if cursor.rowcount:
            member_achievement_id = int(cursor.lastrowid)
            _grant_configured_material_reward(
                conn,
                client_id=client_id,
                definition=definition,
                member_achievement_id=member_achievement_id,
                timezone_name="Europe/Moscow",
            )
            _notify(
                conn,
                client_id=client_id,
                notification_type="achievement",
                title=f"Новое достижение: {definition['name']}",
                body=str(definition["description"] or ""),
                entity_type="achievement",
                entity_id=int(definition["id"]),
            )

    titles = conn.execute(
        "SELECT * FROM title_definitions WHERE is_enabled=1 AND condition_code LIKE 'hijack_%' ORDER BY priority,id"
    ).fetchall()
    for definition in titles:
        condition_code = str(definition["condition_code"])
        value = float(metrics.get(condition_code, 0) or 0)
        if value < float(definition["threshold"]):
            continue
        if str(definition["title_type"]) == "temporary":
            period_key, starts_at, ends_at = _temporary_bounds(condition_code, today, latest_import_id)
            conn.execute(
                """
                INSERT OR IGNORE INTO temporary_title_periods(
                    title_definition_id,period_key,starts_at,ends_at
                ) VALUES (?,?,?,?)
                """,
                (
                    int(definition["id"]),
                    period_key,
                    starts_at.isoformat(timespec="seconds"),
                    ends_at.isoformat(timespec="seconds"),
                ),
            )
            period = conn.execute(
                "SELECT id FROM temporary_title_periods WHERE title_definition_id=? AND period_key=?",
                (int(definition["id"]), period_key),
            ).fetchone()
            temporary_period_id = int(period["id"])
            existing = conn.execute(
                """
                SELECT id FROM member_titles
                WHERE client_id=? AND title_definition_id=? AND temporary_period_id=?
                """,
                (client_id, int(definition["id"]), temporary_period_id),
            ).fetchone()
            if existing:
                continue
            cursor = conn.execute(
                """
                INSERT INTO member_titles(
                    client_id,title_definition_id,temporary_period_id,expires_at,source_json
                ) VALUES (?,?,?,?,?)
                """,
                (
                    client_id,
                    int(definition["id"]),
                    temporary_period_id,
                    ends_at.isoformat(timespec="seconds"),
                    json.dumps({"value": value, "threshold": int(definition["threshold"]), "source": "hi_jack"}, ensure_ascii=False),
                ),
            )
        else:
            existing = conn.execute(
                """
                SELECT id FROM member_titles
                WHERE client_id=? AND title_definition_id=? AND temporary_period_id IS NULL
                """,
                (client_id, int(definition["id"])),
            ).fetchone()
            if existing:
                continue
            cursor = conn.execute(
                """
                INSERT INTO member_titles(client_id,title_definition_id,source_json)
                VALUES (?,?,?)
                """,
                (
                    client_id,
                    int(definition["id"]),
                    json.dumps({"value": value, "threshold": int(definition["threshold"]), "source": "hi_jack"}, ensure_ascii=False),
                ),
            )
        member_title_id = int(cursor.lastrowid)
        _grant_configured_material_reward(
            conn,
            client_id=client_id,
            definition=definition,
            member_title_id=member_title_id,
            timezone_name="Europe/Moscow",
        )
        _notify(
            conn,
            client_id=client_id,
            notification_type="title",
            title=f"Новое звание: {definition['name']}",
            body=str(definition["description"] or ""),
            entity_type="member_title",
            entity_id=member_title_id,
        )
    return metrics


def referral_tree(
    conn: sqlite3.Connection,
    *,
    root_client_id: int,
    max_depth: int = MAX_REFERRAL_TREE_DEPTH,
) -> dict[str, Any]:
    rows = conn.execute(
        """
        SELECT rqp.referrer_client_id,rqp.invited_client_id,
               rqp.distinct_completed_days,rqp.qualified_at,
               c.nickname,c.first_name,c.username
        FROM referral_qualification_progress rqp
        JOIN clients c ON c.id=rqp.invited_client_id
        ORDER BY rqp.created_at,rqp.id
        """
    ).fetchall()
    children: dict[int, list[sqlite3.Row]] = defaultdict(list)
    for row in rows:
        children[int(row["referrer_client_id"])].append(row)
    root = conn.execute(
        "SELECT id,nickname,first_name,username FROM clients WHERE id=?",
        (root_client_id,),
    ).fetchone()
    if not root:
        return {"root": None, "direct": 0, "total": 0, "max_depth": 0, "truncated": False}

    truncated = False
    deepest = 0
    total = 0

    def build(client_id: int, depth: int, path: set[int]) -> list[dict[str, Any]]:
        nonlocal truncated, deepest, total
        if depth > max_depth:
            truncated = True
            return []
        result: list[dict[str, Any]] = []
        for row in children.get(client_id, []):
            invited_id = int(row["invited_client_id"])
            if invited_id in path:
                truncated = True
                continue
            total += 1
            deepest = max(deepest, depth)
            node = {
                "client_id": invited_id,
                "display_name": _display_name({**dict(row), "client_id": invited_id}),
                "qualified": bool(row["qualified_at"]),
                "completed_days": int(row["distinct_completed_days"] or 0),
                "depth": depth,
            }
            node["children"] = build(invited_id, depth + 1, path | {invited_id})
            result.append(node)
        return result

    root_payload = {
        "client_id": int(root["id"]),
        "display_name": _display_name({**dict(root), "client_id": int(root["id"])}),
        "depth": 0,
        "children": build(int(root["id"]), 1, {int(root["id"])}),
    }
    return {
        "root": root_payload,
        "direct": len(root_payload["children"]),
        "total": total,
        "max_depth": deepest,
        "truncated": truncated,
    }
